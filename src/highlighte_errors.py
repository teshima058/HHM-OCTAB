import os
import pandas as pd
import openpyxl as xl
from openpyxl.styles import PatternFill

annotation_file = "./data/annotation_results_integrated_20210528.xlsx"
save_path = "./data/annotation_results_integrated_20210528_mark.xlsx"

annot_df = pd.read_excel(annotation_file)

dic = {'Gesture ID':[], 'Text':[], 'Remark':[]}
exclude_cnt = 0
missed_cols = []
for i in range(len(annot_df)):
    if pd.isna(annot_df['Remarks'].iloc[i]) or pd.isna(annot_df['Text'].iloc[i]):
        continue

    text = str.lower(annot_df['Text'].iloc[i])
    remark = str.lower(annot_df['Remarks'].iloc[i])

    if not remark in text:
        missed_cols.append(i)

print('Mistaked', len(missed_cols), '/', len(annot_df))

# read input xlsx
wb1 = xl.load_workbook(filename=annotation_file)
ws1 = wb1.worksheets[0]

# set gray color cell fill
yellow = PatternFill(patternType='solid', fgColor='ffff00')
white = PatternFill(patternType='solid', fgColor='ffffff')

# write in sheet
for i,row in enumerate(ws1):
    if i == 0:
        continue
    if i - 1 in missed_cols:
        for cell in row:
            ws1[cell.coordinate].fill = yellow
    else:
        for cell in row:
            ws1[cell.coordinate].fill = white

# save target xlsx file
wb1.save(save_path)

print("Saved to {}".format(save_path))

